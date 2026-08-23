"""
app.py — Smart Student Tracking System

Flask application: admin/driver/parent auth, student & bus management,
face-recognition attendance triggers, GPS tracking endpoints, email
notifications, and Excel/PDF reporting.

Changes from the previous version (see chat for full explanation):
- All routes now use database.get_db_connection() (pooled) instead of a
  mix of a stale global connection and ad-hoc hardcoded connections.
- Duplicate imports and duplicate app.secret_key assignment removed.
- Secret key / mail credentials read from environment variables, with
  fallbacks equal to your previous values so behavior is unchanged.
- No routes, URLs, templates, or query logic were removed or renamed.
"""

import os
import subprocess
import sys
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, jsonify
)
from flask_mail import Mail, Message
import openpyxl
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from database import get_db_connection

app = Flask(__name__)

# --- Core config -------------------------------------------------------
# Prefer environment variables in production; fallbacks keep this running
# out of the box exactly as before.
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "smart_student_tracking")

app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME", "")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD", "")
app.config["MAIL_DEFAULT_SENDER"] = os.environ.get("MAIL_DEFAULT_SENDER", "")

mail = Mail(app)


# =========================================================================
# Public / Auth
# =========================================================================

@app.route("/")
def home():
    return render_template("index.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT * FROM users WHERE username=%s AND password=%s",
                (username, password)
            )
            user = cursor.fetchone()
            cursor.close()
        finally:
            conn.close()

        if user:
            session["user_id"] = user["user_id"]
            session["username"] = user["username"]
            session["role"] = user["role"]
            session["reference_id"] = user["reference_id"]

            if user["role"] == "admin":
                return redirect(url_for("dashboard"))
            elif user["role"] == "driver":
                return redirect(url_for("driver_dashboard"))
            elif user["role"] == "parent":
                return redirect(url_for("parent_dashboard"))

        flash("Invalid Username or Password")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/admin_login")
def admin_login():
    return render_template("login.html")


@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/contact")
def contact():
    return render_template("contact.html")


# =========================================================================
# Students
# =========================================================================

@app.route("/students")
def students():
    return render_template("students.html")


@app.route("/add_student", methods=["GET"])
def add_student_page():
    return render_template("students.html")


@app.route("/add_student", methods=["POST"])
def add_student():
    student_name = request.form["student_name"]
    roll_no = request.form["roll_no"]
    department = request.form["department"]
    year = request.form["year"]
    bus_no = request.form["bus_no"]
    parent_name = request.form["parent_name"]
    parent_phone = request.form["parent_phone"]

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO students
            (student_name, roll_no, department, year, bus_no, parent_name, parent_phone)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (student_name, roll_no, department, year, bus_no, parent_name, parent_phone))
        conn.commit()
        cursor.close()
    finally:
        conn.close()

    return "Student Added Successfully!"


@app.route("/view_students")
def view_students():
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT students.*, buses.bus_no, buses.route_name
            FROM students
            LEFT JOIN buses ON students.bus_id = buses.bus_id
        """)
        students_list = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()

    return render_template("view_students.html", students=students_list)


@app.route("/delete_student/<int:id>")
def delete_student(id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM students WHERE student_id=%s", (id,))
        conn.commit()
        cursor.close()
    finally:
        conn.close()

    return redirect(url_for("view_students"))


@app.route("/edit_student/<int:id>")
def edit_student(id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM students WHERE student_id=%s", (id,))
        student = cursor.fetchone()
        cursor.close()
    finally:
        conn.close()

    return render_template("edit_student.html", student=student)


@app.route("/update_student/<int:id>", methods=["POST"])
def update_student(id):
    student_name = request.form["student_name"]
    roll_no = request.form["roll_no"]
    department = request.form["department"]
    year = request.form["year"]
    bus_no = request.form["bus_no"]
    parent_name = request.form["parent_name"]
    parent_phone = request.form["parent_phone"]

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE students
            SET student_name=%s, roll_no=%s, department=%s, year=%s,
                bus_no=%s, parent_name=%s, parent_phone=%s
            WHERE student_id=%s
        """, (student_name, roll_no, department, year, bus_no,
              parent_name, parent_phone, id))
        conn.commit()
        cursor.close()
    finally:
        conn.close()

    return redirect(url_for("view_students"))


@app.route("/assign_bus/<int:id>", methods=["GET", "POST"])
def assign_bus(id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)

        if request.method == "POST":
            bus_id = request.form["bus_id"]
            cursor.execute(
                "UPDATE students SET bus_id=%s WHERE student_id=%s",
                (bus_id, id)
            )
            conn.commit()
            cursor.close()
            return redirect(url_for("view_students"))

        cursor.execute("SELECT * FROM students WHERE student_id=%s", (id,))
        student = cursor.fetchone()

        cursor.execute("SELECT * FROM buses")
        buses = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()

    return render_template("assign_bus.html", student=student, buses=buses)


# =========================================================================
# Buses
# =========================================================================

@app.route("/add_bus", methods=["GET", "POST"])
def add_bus():
    if request.method == "POST":
        bus_no = request.form["bus_no"]
        driver_name = request.form["driver_name"]
        driver_phone = request.form["driver_phone"]
        route = request.form["route"]
        capacity = request.form["capacity"]

        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO buses (bus_no, driver_name, driver_phone, route_name, capacity)
                VALUES (%s, %s, %s, %s, %s)
            """, (bus_no, driver_name, driver_phone, route, capacity))
            conn.commit()
            cursor.close()
        finally:
            conn.close()

        return redirect(url_for("dashboard"))

    return render_template("add_bus.html")


@app.route("/view_buses")
def view_buses():
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM buses")
        buses = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()

    return render_template("view_buses.html", buses=buses)


@app.route("/edit_bus/<int:id>", methods=["GET", "POST"])
def edit_bus(id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)

        if request.method == "POST":
            bus_no = request.form["bus_no"]
            driver_name = request.form["driver_name"]
            driver_phone = request.form["driver_phone"]
            route_name = request.form["route_name"]
            capacity = request.form["capacity"]

            cursor.execute("""
                UPDATE buses
                SET bus_no=%s, driver_name=%s, driver_phone=%s,
                    route_name=%s, capacity=%s
                WHERE bus_id=%s
            """, (bus_no, driver_name, driver_phone, route_name, capacity, id))
            conn.commit()
            cursor.close()
            return redirect(url_for("view_buses"))

        cursor.execute("SELECT * FROM buses WHERE bus_id=%s", (id,))
        bus = cursor.fetchone()
        cursor.close()
    finally:
        conn.close()

    return render_template("edit_bus.html", bus=bus)


@app.route("/delete_bus/<int:id>")
def delete_bus(id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM buses WHERE bus_id=%s", (id,))
        conn.commit()
        cursor.close()
    finally:
        conn.close()

    return redirect(url_for("view_buses"))


# =========================================================================
# Face Recognition triggers
# =========================================================================

@app.route("/start_recognition")
def start_recognition():
    subprocess.Popen([sys.executable, "recognize_face.py"])
    return redirect(url_for("dashboard"))


@app.route("/capture_faces")
def capture_faces():
    subprocess.Popen([sys.executable, "face_capture.py"])
    flash("Face capture started.")
    return redirect(url_for("students"))


@app.route("/train_model")
def train_model():
    subprocess.Popen([sys.executable, "train_model.py"])
    flash("Training started.")
    return redirect(url_for("dashboard"))


# =========================================================================
# Attendance
# =========================================================================

@app.route("/attendance")
def attendance():
    search = request.args.get("search", "")
    date = request.args.get("date", "")

    query = """
        SELECT
            a.attendance_id, s.student_name, s.roll_no,
            b.bus_no, b.route_name,
            a.attendance_date, a.attendance_time, a.status
        FROM attendance a
        JOIN students s ON a.student_id = s.student_id
        LEFT JOIN buses b ON s.bus_id = b.bus_id
        WHERE 1=1
    """
    values = []

    if search:
        query += " AND (s.student_name LIKE %s OR s.roll_no LIKE %s)"
        values += [f"%{search}%", f"%{search}%"]

    if date:
        query += " AND a.attendance_date=%s"
        values.append(date)

    query += " ORDER BY a.attendance_date DESC, a.attendance_time DESC"

    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute(query, values)
        records = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()

    return render_template("attendance.html", records=records, search=search, date=date)


def _fetch_all_attendance_records():
    """Shared query for the Excel and PDF exporters."""
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT
                s.student_name, s.roll_no, b.bus_no, b.route_name,
                a.attendance_date, a.attendance_time, a.status
            FROM attendance a
            JOIN students s ON a.student_id = s.student_id
            LEFT JOIN buses b ON s.bus_id = b.bus_id
            ORDER BY a.attendance_date DESC, a.attendance_time DESC
        """)
        records = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()
    return records


@app.route("/export_excel")
def export_excel():
    records = _fetch_all_attendance_records()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"

    headers = ["Student Name", "Roll No", "Bus No", "Route", "Date", "Time", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    row_num = 2
    for row in records:
        sheet.cell(row=row_num, column=1).value = row["student_name"]
        sheet.cell(row=row_num, column=2).value = row["roll_no"]
        sheet.cell(row=row_num, column=3).value = row["bus_no"]
        sheet.cell(row=row_num, column=4).value = row["route_name"]
        sheet.cell(row=row_num, column=5).value = str(row["attendance_date"])
        sheet.cell(row=row_num, column=6).value = str(row["attendance_time"])
        sheet.cell(row=row_num, column=7).value = row["status"]
        row_num += 1

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Attendance_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/export_pdf")
def export_pdf():
    records = _fetch_all_attendance_records()

    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=letter)

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(180, 780, "Attendance Report")

    y = 750
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(30, y, "Name")
    pdf.drawString(150, y, "Roll No")
    pdf.drawString(250, y, "Bus")
    pdf.drawString(320, y, "Date")
    pdf.drawString(400, y, "Time")
    pdf.drawString(480, y, "Status")
    y -= 20

    pdf.setFont("Helvetica", 9)
    for row in records:
        if y < 40:
            pdf.showPage()
            y = 780

        pdf.drawString(30, y, str(row["student_name"]))
        pdf.drawString(150, y, str(row["roll_no"]))
        pdf.drawString(250, y, str(row["bus_no"] or "-"))
        pdf.drawString(320, y, str(row["attendance_date"]))
        pdf.drawString(400, y, str(row["attendance_time"]))
        pdf.drawString(480, y, str(row["status"]))
        y -= 20

    pdf.save()
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="Attendance_Report.pdf",
        mimetype="application/pdf"
    )


# =========================================================================
# Dashboard
# =========================================================================

@app.route("/dashboard")
def dashboard():
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT COUNT(*) AS total FROM students")
        total_students = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) AS total FROM buses")
        total_buses = cursor.fetchone()["total"]

        cursor.execute("""
            SELECT COUNT(*) AS total FROM attendance
            WHERE attendance_date = CURDATE()
        """)
        today_attendance = cursor.fetchone()["total"]

        cursor.execute("""
            SELECT COUNT(*) AS total FROM attendance
            WHERE status='Present' AND attendance_date = CURDATE()
        """)
        present = cursor.fetchone()["total"]

        cursor.execute("""
            SELECT COUNT(*) AS total FROM attendance
            WHERE status='Absent' AND attendance_date = CURDATE()
        """)
        absent = cursor.fetchone()["total"]

        cursor.execute("""
            SELECT s.student_name, s.roll_no, s.bus_no, a.status
            FROM attendance a
            INNER JOIN students s ON a.student_id = s.student_id
            ORDER BY a.attendance_id DESC
            LIMIT 10
        """)
        recent_attendance = cursor.fetchall()

        cursor.execute("""
            SELECT attendance_date, COUNT(*) AS total
            FROM attendance
            GROUP BY attendance_date
            ORDER BY attendance_date DESC
            LIMIT 7
        """)
        chart = cursor.fetchall()
        chart.reverse()

        cursor.close()
    finally:
        conn.close()

    labels = [str(r["attendance_date"]) for r in chart]
    values = [r["total"] for r in chart]

    return render_template(
        "dashboard.html",
        total_students=total_students,
        total_buses=total_buses,
        today_attendance=today_attendance,
        present=present,
        absent=absent,
        recent_attendance=recent_attendance,
        labels=labels,
        values=values
    )


@app.route("/admin_dashboard")
def admin_dashboard():
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT COUNT(*) AS total FROM students")
        total_students = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) AS total FROM drivers")
        total_drivers = cursor.fetchone()["total"]

        cursor.execute("""
            SELECT COUNT(*) AS total FROM attendance
            WHERE attendance_date = CURDATE() AND status='Present'
        """)
        present_today = cursor.fetchone()["total"]

        absent_today = total_students - present_today
        cursor.close()
    finally:
        conn.close()

    return render_template(
        "admin_dashboard.html",
        total_students=total_students,
        total_drivers=total_drivers,
        present_today=present_today,
        absent_today=absent_today
    )


# =========================================================================
# GPS / Location
# =========================================================================

@app.route("/update_location/<int:bus_id>", methods=["POST"])
def update_location(bus_id):
    latitude = request.form["latitude"]
    longitude = request.form["longitude"]

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE buses SET latitude=%s, longitude=%s WHERE bus_id=%s
        """, (latitude, longitude, bus_id))
        conn.commit()
        cursor.close()
    finally:
        conn.close()

    return "Location Updated"


@app.route("/bus_location/<int:bus_id>")
def bus_location(bus_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT latitude, longitude FROM buses WHERE bus_id = %s
        """, (bus_id,))
        bus = cursor.fetchone()
        cursor.close()
    finally:
        conn.close()

    if bus is None:
        return f"No bus found with bus_id = {bus_id}"

    return render_template("bus_location.html", latitude=bus["latitude"], longitude=bus["longitude"])


@app.route("/get_bus_location/<int:bus_id>")
def get_bus_location(bus_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT latitude, longitude FROM buses WHERE bus_id=%s
        """, (bus_id,))
        row = cursor.fetchone()
        cursor.close()
    finally:
        conn.close()

    if row:
        return jsonify({"latitude": row["latitude"], "longitude": row["longitude"]})

    return jsonify({"error": "Bus not found"})


@app.route("/update_bus_location", methods=["POST"])
def update_bus_location():
    """
    Used by driver_dashboard.html's browser-geolocation script, and by the
    ESP32 + NEO-6M module once wired up.

    FIX: this previously wrote to a `bus_location` table that nothing else
    read from, while /get_bus_location and the parent/admin maps read from
    `buses.latitude/longitude` — so a driver's live position never actually
    reached anyone's screen. Now targets `buses`, same as the rest of the
    GPS pipeline.
    """
    latitude = request.form.get("latitude")
    longitude = request.form.get("longitude")
    bus_id = request.form.get("bus_id")

    if not all([latitude, longitude, bus_id]):
        return jsonify({"error": "latitude, longitude and bus_id are required"}), 400

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE buses
            SET latitude=%s, longitude=%s
            WHERE bus_id=%s
        """, (latitude, longitude, bus_id))
        conn.commit()
        cursor.close()
    finally:
        conn.close()

    return jsonify({"message": "Location updated"})


# =========================================================================
# Driver
# =========================================================================

@app.route("/driver_login", methods=["GET", "POST"])
def driver_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT * FROM drivers WHERE username=%s AND password=%s
            """, (username, password))
            driver = cursor.fetchone()
            cursor.close()
        finally:
            conn.close()

        if driver:
            session["driver_id"] = driver["driver_id"]
            session["bus_id"] = driver["bus_id"]
            return redirect("/driver_dashboard")

        return "Invalid Username or Password"

    return render_template("driver_login.html")


@app.route("/driver_dashboard")
def driver_dashboard():
    return render_template("driver_dashboard.html")


# =========================================================================
# Parent
# =========================================================================

@app.route("/parent_login", methods=["GET", "POST"])
def parent_login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT * FROM parents WHERE email=%s AND password=%s
            """, (email, password))
            parent = cursor.fetchone()
            cursor.close()
        finally:
            conn.close()

        if parent:
            session["parent_id"] = parent["parent_id"]
            session["student_id"] = parent["student_id"]
            return redirect("/parent_dashboard")

        return "Invalid Email or Password"

    return render_template("parent_login.html")


@app.route("/parent_dashboard")
def parent_dashboard():
    if "student_id" not in session:
        return redirect("/parent_login")

    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT s.student_id, s.student_name, s.roll_no,
                   b.bus_id, b.bus_no, b.route_name
            FROM students s
            LEFT JOIN buses b ON s.bus_id = b.bus_id
            WHERE s.student_id = %s
        """, (session["student_id"],))
        student = cursor.fetchone()

        cursor.execute("""
            SELECT attendance_date, status
            FROM attendance
            WHERE student_id = %s
            ORDER BY attendance_date DESC
        """, (session["student_id"],))
        attendance = cursor.fetchall()

        cursor.close()
    finally:
        conn.close()

    return render_template("parent_dashboard.html", student=student, attendance=attendance)


# =========================================================================
# Email
# =========================================================================

def send_attendance_email(parent_email, student_name):
    msg = Message(
        subject="Student Bus Attendance",
        sender=app.config["MAIL_USERNAME"],
        recipients=[parent_email]
    )
    msg.body = f"""Hello,

Your child {student_name} has been marked PRESENT and has boarded the school bus successfully.

Thank you,
Smart Student Tracking System
"""
    mail.send(msg)


@app.route("/test_email")
def test_email():
    try:
        msg = Message(
            subject="Test Email",
            recipients=["harishthirusangu150707@gmail.com"]
        )

        msg.body = "This is a test email."

        mail.send(msg)
        return "Email Sent Successfully"

    except Exception as e:
        return f"Error: {e}"
@app.route("/add_driver", methods=["GET", "POST"])
def add_driver():
    if request.method == "POST":
        driver_name = request.form["driver_name"]
        phone = request.form["phone"]
        username = request.form["username"]
        password = request.form["password"]
        bus_id = request.form.get("bus_id") or None

        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO drivers (driver_name, phone, username, password, bus_id)
                VALUES (%s, %s, %s, %s, %s)
            """, (driver_name, phone, username, password, bus_id))
            conn.commit()
            cursor.close()
        finally:
            conn.close()

        flash("Driver added successfully.")
        return redirect(url_for("view_drivers"))

    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM buses")
        buses = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()

    return render_template("add_driver.html", buses=buses)


@app.route("/view_drivers")
def view_drivers():
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT d.*, b.bus_no, b.route_name
            FROM drivers d
            LEFT JOIN buses b ON d.bus_id = b.bus_id
        """)
        drivers = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()

    return render_template("view_drivers.html", drivers=drivers)


@app.route("/edit_driver/<int:id>", methods=["GET", "POST"])
def edit_driver(id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)

        if request.method == "POST":
            driver_name = request.form["driver_name"]
            phone = request.form["phone"]
            username = request.form["username"]
            bus_id = request.form.get("bus_id") or None
            new_password = request.form.get("password")

            if new_password:
                cursor.execute("""
                    UPDATE drivers
                    SET driver_name=%s, phone=%s, username=%s, password=%s, bus_id=%s
                    WHERE driver_id=%s
                """, (driver_name, phone, username, new_password, bus_id, id))
            else:
                cursor.execute("""
                    UPDATE drivers
                    SET driver_name=%s, phone=%s, username=%s, bus_id=%s
                    WHERE driver_id=%s
                """, (driver_name, phone, username, bus_id, id))

            conn.commit()
            cursor.close()
            return redirect(url_for("view_drivers"))

        cursor.execute("SELECT * FROM drivers WHERE driver_id=%s", (id,))
        driver = cursor.fetchone()

        cursor.execute("SELECT * FROM buses")
        buses = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()

    return render_template("edit_driver.html", driver=driver, buses=buses)


@app.route("/delete_driver/<int:id>")
def delete_driver(id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM drivers WHERE driver_id=%s", (id,))
        conn.commit()
        cursor.close()
    finally:
        conn.close()

    return redirect(url_for("view_drivers"))


@app.route("/view_parents")
def view_parents():
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.parent_id, p.parent_name, p.email, s.student_name, s.roll_no
            FROM parents p
            JOIN students s ON p.student_id = s.student_id
        """)
        parents = cursor.fetchall()
        cursor.close()
    finally:
        conn.close()

    return render_template("view_parents.html", parents=parents)


@app.route("/delete_parent/<int:id>")
def delete_parent(id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM parents WHERE parent_id=%s", (id,))
        conn.commit()
        cursor.close()
    finally:
        conn.close()

    return redirect(url_for("view_parents"))
@app.route("/parent_register", methods=["GET", "POST"])
def parent_register():
    if request.method == "POST":
        name = request.form.get("name")
        email = request.form.get("email")
        password = request.form.get("password")
        phone = request.form.get("phone")

        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            cursor.execute(
                """
                INSERT INTO parents (name, email, password, phone)
                VALUES (%s, %s, %s, %s)
                """,
                (name, email, password, phone)
            )

            conn.commit()

            return redirect(url_for("parent_login"))

        except Exception as e:
            conn.rollback()
            return f"Registration failed: {e}", 500

        finally:
            cursor.close()
            conn.close()

    return render_template("parent_register.html")


if __name__ == "__main__":
    app.run(debug=True)
